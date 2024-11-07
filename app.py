import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import Workbook, load_workbook
import logging
from datetime import datetime
import streamlit as st
from io import BytesIO
import base64

def setup_logging():
    log_filename = f"etl_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    logging.basicConfig(filename=log_filename, level=logging.INFO,
                        format='%(asctime)s - %(levelname)s - %(message)s')

def process_table_data(df_powerapp, df_source):
    logging.info("Début du traitement de la table DATA")

    # Combiner les colonnes 'KPI' et 'Maille d'analyse' du df_source pour obtenir toutes les DATA possibles
    df_source['KPI'] = df_source['KPI'].astype(str).fillna('')
    df_source["Maille d'analyse"] = df_source["Maille d'analyse"].astype(str).fillna('')

    # Fonction pour extraire les valeurs uniques de KPI et Maille d'analyse
    def extract_unique_data(series):
        return set(
            i.strip().lower()
            for items in series
            for i in items.split(',')
            if i.strip()
        )

    kpi_data = extract_unique_data(df_source['KPI'])
    maille_data = extract_unique_data(df_source["Maille d'analyse"])

    # Créer un DataFrame avec toutes les DATA possibles du fichier source
    all_data = pd.DataFrame({
        'DATA': list(kpi_data.union(maille_data))
    })

    # Nettoyage de la colonne 'DATA' dans df_powerapp
    df_powerapp['DATA'] = df_powerapp['DATA'].astype(str).str.strip().str.lower()

    # Sélection des colonnes à conserver de df_powerapp
    columns_to_keep = ["Descriptif de la donnée", "DATA", "Qualité", "Règles de calcul KPI",
                       "Descriptif KPI", "Lien Wiki", "Famille donnée"]
    dftest = df_powerapp[[col for col in columns_to_keep if col in df_powerapp.columns]].copy()

    # Nettoyage de la colonne 'DATA' dans dftest
    dftest['DATA'] = dftest['DATA'].astype(str).str.strip().str.lower()

    # Fusionner les données pour obtenir toutes les DATA avec leurs informations
    dftest = pd.merge(all_data, dftest, on='DATA', how='left')

    # Attribuer un ID_DATA unique à chaque DATA
    dftest.reset_index(drop=True, inplace=True)
    dftest['id'] = dftest.index + 1
    dftest.insert(0, 'ID_DATA', 'DATA' + dftest['id'].astype(str).str.zfill(4))
    dftest.drop('id', axis=1, inplace=True)

    # Déterminer le Type (KPI ou Maille d'analyse)
    dftest['Type'] = dftest['DATA'].apply(
        lambda x: 'KPI' if x in kpi_data else ('Maille d\'analyse' if x in maille_data else 'Non spécifié')
    )

    logging.info("Fin du traitement de la table DATA")
    return dftest

def process_po_data(df_source):
    logging.info("Début du traitement de la table PO DATA")
    df2 = df_source[['PO Data']].copy()
    df_drop = df2.drop_duplicates().reset_index(drop=True)

    def generate_id(phrase, counter):
        logging.info("Début du traitement des ID de la table PO DATA")
        words = phrase.split()
        initials = ''.join([word[0].upper() for word in words])
        id_number = f"{counter:04d}"
        unique_id = initials + id_number
        return unique_id

    df_drop['ID_PO_DATA'] = [generate_id(row, i+1) for i, row in enumerate(df_drop['PO Data'])]
    df_drop.columns = df_drop.columns.map(str)

    logging.info("Fin du traitement de la table PO DATA")
    return df_drop

def process_prompt_data(df_source):
    df2 = df_source[['Ecran de sélection /prompt ']].copy()

    split_data = df2['Ecran de sélection /prompt '].str.split(',', expand=True)
    split_data.columns = [f'Colonne{i+1}' for i in range(split_data.shape[1])]
    df3 = pd.concat([df2, split_data], axis=1)

    id_vars = ["Ecran de sélection /prompt "]
    df_unpivot = pd.melt(df3, id_vars=id_vars, var_name="Colonne", value_name="Valeur")

    df4 = df_unpivot.dropna(subset=['Valeur']).copy()
    df4['Valeur'] = df4['Valeur'].str.lower()
    df4 = df4.drop_duplicates(subset=['Valeur']).reset_index(drop=True)

    df4 = df4[["Ecran de sélection /prompt ", 'Valeur']].copy()

    df4['id'] = df4.index + 1
    df4.insert(0, 'ID_PROMPT', 'PROMPT' + df4['id'].astype(str).str.zfill(4))
    df4.drop('id', axis=1, inplace=True)

    df4.rename(columns={'Valeur': 'Prompt'}, inplace=True)
    df5 = df4[['ID_PROMPT', 'Prompt']].copy()

    df5.columns = df5.columns.map(str)

    return df5

def process_rapport_prompt(df5, df_source):
    df_Rapport_Prompt = df_source[["Nom du rapport", 'Ecran de sélection /prompt ']].copy()

    split_data = df_Rapport_Prompt['Ecran de sélection /prompt '].str.split(',', expand=True)
    split_data.columns = [f'Colonne{i+1}' for i in range(split_data.shape[1])]
    df_Rapport_Prompt2 = pd.concat([df_Rapport_Prompt, split_data], axis=1)

    id_vars = ["Nom du rapport", "Ecran de sélection /prompt "]
    df_unpivot = pd.melt(df_Rapport_Prompt2, id_vars=id_vars, var_name="Colonne", value_name="Valeur")

    df_Rapport_Prompt3 = df_unpivot.dropna(subset=['Valeur']).copy()
    df_Rapport_Prompt3['Valeur'] = df_Rapport_Prompt3['Valeur'].str.lower()
    df_Rapport_Prompt3.rename(columns={'Valeur': 'Prompt'}, inplace=True)

    merged_df_Prompt = pd.merge(df_Rapport_Prompt3, df5, on='Prompt')

    return merged_df_Prompt

def process_rapport_data(merged_df_Prompt, df_source):
    df_Rapport = df_source
    df_Rapport['DATA_detail'] = df_Rapport["KPI"] + ',' + df_Rapport["Maille d'analyse"]
    
    def generer_id(texte_original, index):
        texte_nettoye = texte_original.translate(str.maketrans("", "", "-_?/()–"))
        mots = texte_nettoye.split()
        mots_filtres = [mot for mot in mots if mot.lower() not in {"en", "par", "et", "des", "a"}]
        premieres_lettres = [mot[0].upper() for mot in mots_filtres]
        id_partiel = ''.join(premieres_lettres)
        id_sans_numero = id_partiel[:3]
        id_complet = id_sans_numero + str(index).zfill(4)
        return id_complet
    
    df_Rapport.insert(0, 'ID_RAPPORT', df_Rapport['Nom du rapport'].apply(lambda x: generer_id(x, df_Rapport.index[df_Rapport['Nom du rapport'] == x].tolist()[0] + 1)))
    
    merged_df_Prompt2 = pd.merge(merged_df_Prompt, df_Rapport, on='Nom du rapport')
    df_Rapport_Prompt = merged_df_Prompt2[["Nom du rapport", "Prompt", "ID_PROMPT", "ID_RAPPORT"]]
    
    df_Rapport_Prompt.columns = df_Rapport_Prompt.columns.map(str)
    
    return df_Rapport

def process_rapport_data_2(df_Rapport, df_data):
    df_rapport_data = df_Rapport.copy()

    df_rapport_data['KPI'] = df_rapport_data['KPI'].astype(str).fillna('')
    df_rapport_data['Maille d\'analyse'] = df_rapport_data['Maille d\'analyse'].astype(str).fillna('')

    df_rapport_data['KPI_list'] = df_rapport_data['KPI'].str.split(',')
    df_rapport_data['Maille_list'] = df_rapport_data['Maille d\'analyse'].str.split(',')

    # Nettoyage des listes
    df_rapport_data['KPI_list'] = df_rapport_data['KPI_list'].apply(lambda x: [i.strip().lower() for i in x if i.strip()])
    df_rapport_data['Maille_list'] = df_rapport_data['Maille_list'].apply(lambda x: [i.strip().lower() for i in x if i.strip()])

    df_rapport_data['DATA_list'] = df_rapport_data['KPI_list'] + df_rapport_data['Maille_list']

    df_rapport_data_exploded = df_rapport_data.explode('DATA_list')

    df_rapport_data_exploded['DATA'] = df_rapport_data_exploded['DATA_list']

    # Nettoyage de DATA
    df_rapport_data_exploded['DATA'] = df_rapport_data_exploded['DATA'].str.strip().str.lower()

    # Filtrage des valeurs non nulles
    df_rapport_data_exploded = df_rapport_data_exploded[df_rapport_data_exploded['DATA'].notnull() & (df_rapport_data_exploded['DATA'] != '')]

    # Nettoyage de df_data
    df_data['DATA'] = df_data['DATA'].astype(str).str.strip().str.lower()

    # Fusion des données
    merged_df_rapport_data = pd.merge(df_rapport_data_exploded, df_data, on='DATA', how='left')

    # Vérification des DATA sans correspondance
    missing_id_data = merged_df_rapport_data[merged_df_rapport_data['ID_DATA'].isna()]
    if not missing_id_data.empty:
        logging.warning("Les DATA suivants n'ont pas de correspondance dans df_data :")
        logging.warning(missing_id_data['DATA'].unique())

    df_Rapport_data_final = merged_df_rapport_data[["Nom du rapport", "ID_RAPPORT", "ID_DATA", "DATA", "Type"]].copy()

    df_Rapport_data_final.columns = df_Rapport_data_final.columns.map(str)

    return df_Rapport_data_final

def process_axe_temps(df_source):
    df_Axe_Temps = df_source[['Axe temps du rapport']].drop_duplicates().reset_index(drop=True)

    def generate_id(phrase, counter):
        words = phrase.split()
        initials = ''.join([word[0].upper() for word in words[:3]])
        id_number = f"{counter:04d}"
        unique_id = initials + id_number
        return unique_id

    df_Axe_Temps['ID_AXE_TEMPS'] = [generate_id(row, i+1) for i, row in enumerate(df_Axe_Temps['Axe temps du rapport'])]

    return df_Axe_Temps

def process_rapport_part2(df_Rapport, df_Axe_Temps, df_po_data):
    merged_Rapport = pd.merge(df_Rapport, df_po_data, on='PO Data', how='left')
    merged_Rapport2 = pd.merge(merged_Rapport, df_Axe_Temps, on='Axe temps du rapport', how='left')

    merged_Rapport2.columns = merged_Rapport2.columns.map(str)

    return merged_Rapport2

# code du type kpi ou maille d'analyse
def process_kpi_and_maille(df_source):
    def ensure_list(x):
        if pd.isna(x):
            return []
        elif isinstance(x, list):
            return x
        elif isinstance(x, str):
            return [i.strip().lower() for i in x.split(',') if i.strip()]
        else:
            return [str(x)]

    kpi_series = df_source['KPI'].apply(ensure_list)
    maille_series = df_source["Maille d'analyse"].apply(ensure_list)

    kpi_list = [(item, 'KPI') for sublist in kpi_series for item in sublist]
    maille_list = [(item, "Maille d'analyse") for sublist in maille_series for item in sublist]

    combined_list = kpi_list + maille_list
    df_final = pd.DataFrame(combined_list, columns=['DATA', 'Type']).drop_duplicates()

    return df_final

def create_excel_with_tables(all_dataframes, filename='fichier_transforme.xlsx'):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Écrire les DataFrames dans le fichier Excel
        for sheet_name, df in all_dataframes.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Accéder au classeur et ajouter les styles de tableau
        workbook = writer.book
        for sheet_name in all_dataframes.keys():
            worksheet = workbook[sheet_name]
            df = all_dataframes[sheet_name]
            # Définir la plage de données (toutes les lignes et colonnes du DataFrame)
            nb_rows = df.shape[0] + 1  # +1 pour l'en-tête
            nb_cols = df.shape[1]
            # Définir la plage de cellules à convertir en tableau
            ref = f"A1:{get_column_letter(nb_cols)}{nb_rows}"
            # Créer le tableau
            table = Table(displayName=f"Table_{sheet_name}", ref=ref)
            # Définir le style du tableau
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            table.tableStyleInfo = style
            # Ajouter le tableau à la feuille de calcul
            worksheet.add_table(table)
    output.seek(0)
    return output

def create_excel_with_tables_from_sheets(sheets, filename='Résultat_Données_Similaires.xlsx'):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Écrire les DataFrames dans le fichier Excel
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Accéder au classeur et ajouter les styles de tableau
        workbook = writer.book
        for sheet_name in sheets.keys():
            worksheet = workbook[sheet_name]
            df = sheets[sheet_name]
            # Définir la plage de données (toutes les lignes et colonnes du DataFrame)
            nb_rows = df.shape[0] + 1  # +1 pour l'en-tête
            nb_cols = df.shape[1]
            # Définir la plage de cellules à convertir en tableau
            ref = f"A1:{get_column_letter(nb_cols)}{nb_rows}"
            # Créer le tableau
            table = Table(displayName=f"Table_{sheet_name}", ref=ref)
            # Définir le style du tableau
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            table.tableStyleInfo = style
            # Ajouter le tableau à la feuille de calcul
            worksheet.add_table(table)
    output.seek(0)
    return output

def main_etl(df_powerapp, df_source):
    df_data = process_table_data(df_powerapp, df_source)
    df_po_data = process_po_data(df_source)
    df5 = process_prompt_data(df_source)
    merged_df_Prompt = process_rapport_prompt(df5, df_source)
    df_Rapport = process_rapport_data(merged_df_Prompt, df_source)
    df_Rapport_data = process_rapport_data_2(df_Rapport, df_data)
    df_Axe_Temps = process_axe_temps(df_source)
    merged_Rapport2 = process_rapport_part2(df_Rapport, df_Axe_Temps, df_po_data)

    all_dataframes = {
        'Table_DATA': df_data,
        'Table_Prompt': df5,
        'Table_PO_DATA': df_po_data,
        'Table_Rapport_Prompt': merged_df_Prompt,
        'Table_Rapport_Data': df_Rapport_data,
        'Table_AxeTemps': df_Axe_Temps,
        'Table_Rapport': merged_Rapport2
    }

    return all_dataframes

def process_similar_data(df_data):
    logging.info("Début du traitement des données similaires")
    
    # Création de la colonne contenant les Données Similaires
    df_data['données similaires'] = df_data.apply(
        lambda row: ', '.join(
            df_data.loc[df_data['Descriptif de la donnée'] == row['Descriptif de la donnée'], 'DATA'].unique()
        ), axis=1
    )
    df_data['données similaires'] = df_data.apply(
        lambda row: ', '.join(
            [DATA for DATA in row['données similaires'].split(', ') if DATA != row['DATA']]
        ), axis=1
    )
    
    logging.info("Fin du traitement des données similaires")
    return df_data

def main():
    st.title('Traitement des Données')

    # Liste des pages
    pages = ['Notice d\'utilisation', 'Processus ETL', 'Données Similaires']

    # Initialiser l'index de la page actuelle dans la session
    if 'page_index' not in st.session_state:
        st.session_state.page_index = 0


    selected_page = st.radio("Navigation", pages, index=st.session_state.page_index)

    # Si l'utilisateur sélectionne une page différente via les onglets
    if selected_page != pages[st.session_state.page_index]:
        st.session_state.page_index = pages.index(selected_page)

    if st.session_state.page_index == 0:
        # Page 1 : Notice d'utilisation
        st.header('Notice d\'utilisation')
        st.write("""
        Bienvenue dans l'outil de traitement des données.
        
        **Comment utiliser cet outil :**
        
        1. **Processus ETL** : Cette étape vous permet de transformer vos fichiers "Powerapp Dictionnaire des données BU Colissimo" et "Source Dictionnaire des données BU Colissimo". Veuillez préparer ces deux fichiers avant de continuer. 
        Trouvable ici : 
        - Fichier Powerapp : https://laposte.sharepoint.com/sites/CatalogueDataBUColissimo/Documents%20partages/Forms/AllItems.aspx?FolderCTID=0x01200022620E9D45DCBC4783431EA5B3F34F1A&id=%2Fsites%2FCatalogueDataBUColissimo%2FDocuments%20partages%2FGeneral%2FPowerAPP
        """)
        
        st.image("Image_PA/powerapp.PNG", caption="Fichier Powerapp", use_column_width=True)
        
        st.write("""
        - Fichier Source : https://laposte.sharepoint.com/sites/CatalogueDataBUColissimo/Documents%20partages/Forms/AllItems.aspx?FolderCTID=0x01200022620E9D45DCBC4783431EA5B3F34F1A&id=%2Fsites%2FCatalogueDataBUColissimo%2FDocuments%20partages%2FGeneral%2FSource%20Catalogue%20Data&viewid=80713eb7%2Da4d4%2D4803%2D93e5%2D315fb27322ed
        """)
        
        st.image("Image_PA/Source.PNG", caption="Fichier Source", use_column_width=True)
        
        st.write("""
        2. **Données Similaires** : Après avoir exécuté le processus ETL, il faut remplir les descriptions des nouvelles données. Ensuite, vous pourrez passer à l’étape de traitement des données similaires.


        3. **Fin** : Une fois toute l’étape 2 terminée, téléchargez le fichier et placez-le dans le dossier suivant : https://laposte.sharepoint.com/sites/CatalogueDataBUColissimo/Documents%20partages/Forms/AllItems.aspx?FolderCTID=0x01200022620E9D45DCBC4783431EA5B3F34F1A&id=%2Fsites%2FCatalogueDataBUColissimo%2FDocuments%20partages%2FGeneral%2FPowerAPP&viewid=80713eb7%2Da4d4%2D4803%2D93e5%2D315fb27322ed
        
        Utilisez les boutons **Suivant** et **Retour** ou les onglets de navigation pour parcourir les différentes étapes.
        """)


    elif st.session_state.page_index == 1:
        # Page 2 : Processus ETL
        st.header('Transformation des fichiers Dictionnaire des données')

        uploaded_files = st.file_uploader("Choisissez les fichiers Excel", type=['xlsx'], accept_multiple_files=True, key="etl_files")

        if uploaded_files and len(uploaded_files) == 2:
            files_dict = {
                "Powerapp Dictionnaire des données BU Colissimo": None,
                "Source Dictionnaire des données BU Colissimo": None
            }

            for uploaded_file in uploaded_files:
                if 'Powerapp' in uploaded_file.name:
                    files_dict["Powerapp Dictionnaire des données BU Colissimo"] = uploaded_file
                elif 'Source' in uploaded_file.name:
                    files_dict["Source Dictionnaire des données BU Colissimo"] = uploaded_file

            if None not in files_dict.values():
                if st.button('Exécuter le processus ETL'):
                    try:
                        df_powerapp = pd.read_excel(files_dict["Powerapp Dictionnaire des données BU Colissimo"], sheet_name='Table DATA')
                        df_source = pd.read_excel(files_dict["Source Dictionnaire des données BU Colissimo"])

                        # Appel de la fonction main_etl pour exécuter le processus ETL complet
                        all_dataframes = main_etl(df_powerapp, df_source)

                        st.success("Le processus ETL a été exécuté avec succès.")

                        # Utiliser la nouvelle fonction pour créer le fichier Excel avec les tableaux formatés
                        output = create_excel_with_tables(all_dataframes)

                        # Bouton de téléchargement
                        st.download_button(
                            label="Télécharger le fichier transformé",
                            data=output,
                            file_name="fichier_transforme.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    except Exception as e:
                        st.error(f"Une erreur s'est produite : {str(e)}")
            else:
                st.error("Les fichiers téléchargés ne correspondent pas aux fichiers requis. Veuillez vérifier les noms des fichiers.")
        else:
            st.info("Veuillez télécharger exactement deux fichiers : 'Powerapp Dictionnaire des données BU Colissimo' et 'Source Dictionnaire des données BU Colissimo'.")

    elif st.session_state.page_index == 2:
        # Page 3 : Données Similaires
        st.header('Traitement des Données Similaires')
        st.write("""
        Une fois l’étape 1 terminée, téléchargez le fichier puis remplissez la description des nouvelles données. Une fois cela terminé, vous pouvez joindre le fichier ci-dessous pour ajouter les données similaires.        """)
        etl_result_file = st.file_uploader("Choisissez le fichier transforme ", type=['xlsx'], key="etl_result_file")

        if etl_result_file:
            if st.button('Traiter les données similaires'):
                try:
                    # Lire toutes les feuilles du fichier Excel
                    xls = pd.ExcelFile(etl_result_file)
                    sheets = {sheet_name: pd.read_excel(xls, sheet_name) for sheet_name in xls.sheet_names}

                    if 'Table_DATA' in sheets:
                        sheets['Table_DATA'] = process_similar_data(sheets['Table_DATA'])
                        st.success("Le traitement des données similaires a été effectué avec succès.")
                    else:
                        st.error("La feuille 'Table_DATA' n'existe pas dans le fichier.")
                        st.stop()

                    # Utiliser la nouvelle fonction pour créer le fichier Excel avec les tableaux formatés
                    output = create_excel_with_tables_from_sheets(sheets)

                    st.download_button(
                        label="Télécharger le résultat",
                        data=output,
                        file_name="Powerapp Dictionnaire des données BU Colissimo.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                except Exception as e:
                    st.error(f"Une erreur s'est produite : {str(e)}")
        else:
            st.info("N’oubliez pas de remplir les descriptions des données avant de télécharger le fichier transforme pour le traitement des données similaires.")



    def on_click_prev():
        if st.session_state.page_index > 0:
            st.session_state.page_index -= 1

    def on_click_next():
        if st.session_state.page_index < len(pages) - 1:
            st.session_state.page_index += 1

    # Afficher les boutons en bas de la page
    col1, col2 = st.columns([1, 1])

    with col1:
        st.button('Retour', on_click=on_click_prev, disabled=(st.session_state.page_index == 0))
    with col2:
        st.button('Suivant', on_click=on_click_next, disabled=(st.session_state.page_index == len(pages) - 1))

if __name__ == "__main__":
    main()
